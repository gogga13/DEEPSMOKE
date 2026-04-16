from openpyxl import load_workbook
from pathlib import Path
path = Path(r"C:\Users\User\Downloads\Каталог (1).xlsx")
wb = load_workbook(path, read_only=False, data_only=True)
ws = wb.active
for r in range(2, ws.max_row + 1):
    category = ws.cell(r, 2).value
    name = ws.cell(r, 3).value
    variants = ws.cell(r, 10).value
    if variants not in (None, ''):
        s = str(variants).strip()
        if s == '|' or ('|' not in s and ';' not in s and len(s.split()) <= 4):
            print('CHECK_ROW', r, '|', name, '|', s)
