from openpyxl import load_workbook
from pathlib import Path
path = Path(r"C:\Users\User\Downloads\Каталог (1).xlsx")
wb = load_workbook(path, read_only=False, data_only=True)
ws = wb.active
print('TITLE', ws.title)
print('MAX_ROW', ws.max_row)
print('MAX_COL', ws.max_column)
headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
print('HEADERS', headers)
for r in range(1, min(ws.max_row, 45)+1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 12)+1)]
    print(r, vals)
